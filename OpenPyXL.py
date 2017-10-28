"""
Excel with openpyxl v. 1.0
Coded by Kuba Siekierzyński (c) 2017
Original code:
https://code.sololearn.com/cGcw92Q6ZUjB/#py

The code demonstrates how to process .xlsx files with openpyxl module for Python. First, the input data is pulled from the input file. Then, the data is processed in Python. The result is then saved in the local output file.

The code will not run in SoloLearn's Code Playground. Please run it from your local IDE.

"""

from openpyxl import load_workbook
from openpyxl import Workbook

data = {}
cols = ['ABCDEFGHIJ']
# 'data' will store cell addresses as keys and their values as dictionary values
# cols will make iterations more feasible

InputFilename = 'input_example.xlsx'
# this is the file we will read and process (change adequately)

OutputFilename = 'output_example.xlsx'
# this is the filename to be created (change adequately)

wb_in = load_workbook(filename=InputFileName)
# the file is loaded...
sheet = wb['Sheet1']
# ...and this is the sheet to be processed

# the code below copies the 10x10 table from the input file to data variable
for row in range(1, 11): # remember that rows start from 1 not 0
    for col in cols:
        data[col+str(row)] = sheet[col+str(row)].value

# ok, we know how to extract data from xls already, now let's write it to another spreadsheet!

wb_out = Workbook()
# object of Workbook class
ws = wb_out.active
# ws is a working sheet...
ws.title = 'Results'
# ...and we set its title to 'Results'

for row_num in range(1, 11): # we'll use another variable to avoid confusion below
    for col in range(1, 11):
        ws.cell(column=col, row=row_num, value=data[cols[col-1]+str(row)])

# That's it, but always remember to save the file!
wb_out.save(filename=OutputFilename)
# This command writes to the output file and closes it. Our job is done :)
